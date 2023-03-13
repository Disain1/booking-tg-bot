import utils.status as status

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

from utils.database import Database

redFill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
yellowFill = PatternFill(start_color='FFBD00', end_color='FFBD00', fill_type='solid')

def getPlacesFile(database: Database):
    wb = load_workbook("excel/new.xlsx")
    ws = wb.active

    result = database.getAll("SELECT row, place, status FROM places")

    s = 0
    for row in ws["D4:AC22"]:
        for cell in row:
            if s == 421:
                break

            if str(cell.value) == "0" or cell.value == None:
                continue

            if result[s][2] == status.BOOKED:
                cell.fill = yellowFill
            elif result[s][2] == status.OCCUPIED:
                cell.fill = redFill
            s += 1

    wb.save("excel/c.xlsx")
    return open("excel/c.xlsx", "rb")


def getUsersFile(database: Database):
    wb = Workbook()
    ws = wb.active

    result = database.getAll("SELECT * FROM users")
    
    i = 0
    for row in result:
        i += 1
        j = 1
        for col in row:
            if j == 3:
                col = "Нет" if col == 0 else "Да"
            
            cell = ws.cell(row = i, column = j)
            cell.value = col
            j += 1

    wb.save("excel/u.xlsx")
    return open("excel/u.xlsx", "rb")